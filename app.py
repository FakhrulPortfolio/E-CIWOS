#region Setup
import time
import pandas as pd
import numpy as np
import threading
import io
from flask import Flask, render_template, request, jsonify, Response, url_for, redirect, send_file, make_response
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import logging
#endregion

#region Configuration
app = Flask(__name__)
logging.basicConfig(level=logging.DEBUG)
progress_Incident_Selenium = 0
progress_WorkOrder_Selenium = 0
lock_Incident_Selenium = threading.Lock()
lock_WorkOrder_Selenium = threading.Lock()
#endregion

#region Do Not Disturb
@app.route('/', methods=['GET', 'POST'])
def index():
    return render_template('index.html')

@app.route('/work', methods=['GET', 'POST'])
def work():
    return render_template('work.html')

@app.route('/run-IncidentScript', methods=['POST'])
def run_IncidentScript():
    data = request.json
    email = data.get('email')
    password = data.get('password')
    if email and password:
        run_selenium_IncidentScript(email,password)
        return jsonify({"status": "success", "message": "Task started"})
    return jsonify({"status": "error", "message": "No link provided"}), 400

@app.route('/progress_Incident_Selenium')
def progress_stream_Incident_Selenium():
    def generate():
        global progress_Incident_Selenium
        while progress_Incident_Selenium < 100:
            with lock_Incident_Selenium:
                yield f"data:{progress_Incident_Selenium}\n\n"
            time.sleep(1)
        yield f"data:{progress_Incident_Selenium}\n\n"  # Ensure it sends 100% at the end

    return Response(generate(), mimetype='text/event-stream')

@app.route('/run-WorkOrderScript', methods=['POST'])
def run_WorkOrderScript():
    data = request.json
    email = data.get('email')
    password = data.get('password')
    if email and password:
        run_selenium_WorkOrderScript(email,password)
        return jsonify({"status": "success", "message": "Task started"})
    return jsonify({"status": "error", "message": "No link provided"}), 400

@app.route('/progress_WorkOrder_Selenium')
def progress_stream_WorkOrder_Selenium():
    def generate():
        global progress_WorkOrder_Selenium
        while progress_WorkOrder_Selenium < 100:
            with lock_WorkOrder_Selenium:
                yield f"data:{progress_WorkOrder_Selenium}\n\n"
            time.sleep(1)
        yield f"data:{progress_WorkOrder_Selenium}\n\n"  # Ensure it sends 100% at the end

    return Response(generate(), mimetype='text/event-stream')
#endregion

#region Extracting Incident
def run_selenium_IncidentScript(email, password):

    global progress_Incident_Selenium
    with lock_Incident_Selenium:
        progress_Incident_Selenium = 0

    # Get the website, email and password
    website = 'https://mygenieplus-smartit.onbmc.com/smartit/app/#/ticket-console'

    # Get the first day and last day of the previous month
    start_of_current_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_of_previous_month = start_of_current_month - timedelta(days=1)
    end_of_previous_month = end_of_previous_month.replace(hour=23, minute=59, second=59, microsecond=999999)
    start_of_previous_month = end_of_previous_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Format the dates
    date_format = "%d-%m-%Y %H:%M:%S"
    start_of_previous_month_formatted = start_of_previous_month.strftime(date_format)
    end_of_previous_month_formatted = end_of_previous_month.strftime(date_format)

    # Setup Chrome Driver
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_experimental_option("detach", True)
    service = Service()
    driver = webdriver.Chrome(service=service, options=options)
    driver.get(website)

    time.sleep(5)
    driver.implicitly_wait(10)

    # Input Email
    driver.find_element(By.ID, 'i0116').send_keys(email)
    driver.find_element(By.ID, 'idSIButton9').click()

    with lock_Incident_Selenium:
        progress_Incident_Selenium = 10
    time.sleep(5)

    # Input Password
    driver.find_element(By.ID, 'i0118').send_keys(password)
    driver.find_element(By.ID, 'idSIButton9').click()

    with lock_Incident_Selenium:
        progress_Incident_Selenium = 20
    time.sleep(25)

    # Click 'More' Report
    driver.find_element(By.XPATH, '//a[@ux-id="more-dropdown"]').click()
    time.sleep(10)
    driver.find_element(By.XPATH, '//li[@ux-id="navitem-more_reports"]').click()
    time.sleep(10)

    with lock_Incident_Selenium:
        progress_Incident_Selenium = 30

    # Insert and Find the Report Name
    name_report = 'A. Incident Raw Data (SLA Response + SLA Resolution)'
    driver.switch_to.window(driver.window_handles[1])
    driver.implicitly_wait(80)
    driver.find_element(By.XPATH, '//div[@class="searchBar searchBarExtraOptions"]/input').send_keys(name_report)
    driver.find_element(By.XPATH, '//div[@class="searchImgContainer"]').click()
    time.sleep(20)
    action = ActionChains(driver)
    element = driver.find_element(By.XPATH, f'//div[@class="browseContentItem"]/div[@title="{name_report}"]/parent::div')
    action.double_click(element).perform()

    with lock_Incident_Selenium:
        progress_Incident_Selenium = 40

    try:
        driver.find_element(By.XPATH, '//*[@id="submit"]//div[text()="Submit"]').click()
    except:
        WebDriverWait(driver, 100).until(EC.invisibility_of_element_located((By.XPATH, '//div[@id="submit"]/div/div')))
    
    with lock_Incident_Selenium:
        progress_Incident_Selenium = 50    
    driver.implicitly_wait(180) #time that need to be adjust

    # Delete items in Product Categorization 2 column
    for i in range(100):
        driver.find_element(By.XPATH, '//div[@id="1966929"]//div[@class="removeListItem"]').click()
 
    with lock_Incident_Selenium:
        progress_Incident_Selenium = 60
    driver.implicitly_wait(60) #time that need to adjust

    # Put all product item into Product Categorization 2 column
    dp = pd.read_excel("Dev Tribe Incident & SR Data - Template.xlsx", sheet_name='Squad List 2', keep_default_na=False)

    for prod in dp["CI Name"]:
        driver.find_element(By.XPATH, '(//div[@id="1966929"]//div[@class="removeListItem"]//following-sibling::div/input)[last()]').send_keys(prod)
        driver.find_element(By.XPATH, '(//div[@id="1966929"]//div[@class="addItemView"]//following-sibling::div/input)[last()]').click()

    with lock_Incident_Selenium:
        progress_Incident_Selenium = 70

    driver.find_element(By.XPATH, '//div[@id="1966922"]//div[@class="filterOutput"]//img').click()
    driver.find_element(By.XPATH, '(//div[@class="dateDisplay"]//input)[last()-1]').clear()
    time.sleep(1)
    driver.find_element(By.XPATH, '(//div[@class="dateDisplay"]//input)[last()-1]').send_keys(start_of_previous_month_formatted)
    time.sleep(1)
    driver.find_element(By.XPATH, '(//div[@class="dateDisplay"]//input)[last()]').clear()
    time.sleep(1)
    driver.find_element(By.XPATH, '(//div[@class="dateDisplay"]//input)[last()]').send_keys(end_of_previous_month_formatted)
    time.sleep(1)
    driver.find_element(By.XPATH, '(//div[@class="applyButton"])[last()]').click()

    driver.find_element(By.XPATH, '//td[@class="submitMidHighlightButton"]/span[text()="Go"]').click()

    with lock_Incident_Selenium:
        progress_Incident_Selenium = 80

    try:
        driver.find_element(By.XPATH, '//*[@id="submit"]//div[text()="Submit"]').click()
    except:
        WebDriverWait(driver, 100).until(EC.invisibility_of_element_located((By.XPATH, '//div[@id="submit"]/div/div')))

    with lock_Incident_Selenium:
        progress_Incident_Selenium = 90
    time.sleep(350)

    # Export the report
    driver.find_element(By.XPATH, '//td[@id="reportexport"]').click()
    driver.find_element(By.XPATH, '//div[@id="rptDataOverlayPanelContent"]//tbody/tr[last()-1]').click()
    driver.find_element(By.XPATH, '//div[@id="xlsExportBtnContainer"]/button').click()
    
    with lock_Incident_Selenium:
        progress_Incident_Selenium = 100
    time.sleep(100)
    #driver.quit()
#endregion

#region Extracting Work Order
def run_selenium_WorkOrderScript(email, password):
    global progress_WorkOrder_Selenium
    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 0

    # Fill your email and password below
    website = 'https://mygenieplus-smartit.onbmc.com/smartit/app/#/ticket-console'

    # Get the first day of the current month
    start_of_current_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_of_previous_month = start_of_current_month - timedelta(days=1)
    end_of_previous_month = end_of_previous_month.replace(hour=23, minute=59, second=59, microsecond=999999)
    start_of_previous_month = end_of_previous_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Format the dates
    date_format = "%d-%m-%Y %H:%M:%S"
    start_of_previous_month_formatted = start_of_previous_month.strftime(date_format)
    end_of_previous_month_formatted = end_of_previous_month.strftime(date_format)

    # Config
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_experimental_option("detach", True)
    service = Service()
    driver = webdriver.Chrome(service=service, options=options)

    # Login 
    driver.get(website)
    time.sleep(5)

    driver.implicitly_wait(10)
    driver.find_element(By.ID, 'i0116').send_keys(email)
    driver.find_element(By.ID, 'idSIButton9').click()
    time.sleep(5)
    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 10
    print("Progress is 10%")

    driver.find_element(By.ID, 'i0118').send_keys(password)
    driver.find_element(By.ID, 'idSIButton9').click()
    time.sleep(25)
    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 20
    print("Progress is 20%")

    # driver.find_element(By.ID, 'idBtn_Back').click()

    # Filter by Folder
    # driver.get(website)

    driver.find_element(By.XPATH, '//a[@ux-id="more-dropdown"]').click()
    time.sleep(50)
    driver.find_element(By.XPATH, '//li[@ux-id="navitem-more_reports"]').click()
    time.sleep(30)

    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 30


    # Search by Name
    name_report = 'A. SR + WO - Raw Data All'
    driver.switch_to.window(driver.window_handles[1])
    driver.implicitly_wait(30)
    driver.find_element(By.XPATH, '//div[@class="searchBar searchBarExtraOptions"]/input').send_keys(name_report)
    driver.find_element(By.XPATH, '//div[@class="searchImgContainer"]').click()
    # Filter in report
    time.sleep(10)
    action = ActionChains(driver)
    element = driver.find_element(By.XPATH, f'//div[@class="browseContentItem"]/div[@title="{name_report}"]/parent::div')
    action.double_click(element).perform()
    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 40
    print("Progress is 40%")

    try:
        driver.find_element(By.XPATH, '//*[@id="submit"]//div[text()="Submit"]').click()
    except:
        WebDriverWait(driver, 100).until(EC.invisibility_of_element_located((By.XPATH, '//div[@id="submit"]/div/div')))

    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 50
    driver.implicitly_wait(60)

    # changing id here
    for i in range(6):
        driver.find_element(By.XPATH, '//div[@id="1952396"]//div[@class="removeListItem"]').click()
    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 60
    print("Progress is 60%")
    driver.implicitly_wait(60)

    items_to_search = [
        "PD UPSTREAM-DEVELOPMENT",
        "PD UPSTREAM-EXPLORATION",
        "PD UPSTREAM-REGULATOR",
        "PD UPSTREAM-PRODUCTION",
        "PD UPSTREAM-IT INTERNATIONAL",
        "PD UPSTREAM-IT MALAYSIA"
    ]

    # changing id here
    for prod in items_to_search:
        driver.find_element(By.XPATH, '(//div[@id="1952396"]//div[@class="removeListItem"]//following-sibling::div/input)[last()]').send_keys(prod)
        driver.find_element(By.XPATH, '(//div[@id="1952396"]//div[@class="addItemView"]//following-sibling::div/input)[last()]').click()
    
    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 70

    # changing id here
    # Get previous date
    driver.find_element(By.XPATH, '//div[@id="1952407"]//div[@class="filterOutput"]//img').click()
    driver.find_element(By.XPATH, '(//div[@class="dateDisplay"]//input)[last()-1]').clear()
    time.sleep(1)
    driver.find_element(By.XPATH, '(//div[@class="dateDisplay"]//input)[last()-1]').send_keys(start_of_previous_month_formatted)
    time.sleep(1)
    driver.find_element(By.XPATH, '(//div[@class="dateDisplay"]//input)[last()]').clear()
    time.sleep(1)
    driver.find_element(By.XPATH, '(//div[@class="dateDisplay"]//input)[last()]').send_keys(end_of_previous_month_formatted)
    time.sleep(1)
    driver.find_element(By.XPATH, '(//div[@class="applyButton"])[last()]').click()
    
    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 80
    print("Progress is 80%")

    driver.find_element(By.XPATH, '//td[@class="submitMidHighlightButton"]/span[text()="Go"]').click()
    try:
        driver.find_element(By.XPATH, '//*[@id="submit"]//div[text()="Submit"]').click()
    except:
        WebDriverWait(driver, 100).until(EC.invisibility_of_element_located((By.XPATH, '//div[@id="submit"]/div/div')))
    
    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 90
    print("Progress is 90%")
    time.sleep(60)
    # Extract report
    driver.find_element(By.XPATH, '//td[@id="reportexport"]').click()
    driver.find_element(By.XPATH, '//div[@id="rptDataOverlayPanelContent"]//tbody/tr[last()-1]').click()
    driver.find_element(By.XPATH, '//div[@id="xlsExportBtnContainer"]/button').click()
    
    with lock_WorkOrder_Selenium:
        progress_WorkOrder_Selenium = 100
    print("Progress is 100%")
    #driver.quit()
#endregion

#region Cleaning Incident
@app.route('/process-excel-Incident', methods=['POST'])
def process_excel_Incident():

    try:
        # Get the uploaded file
        file = request.files['file']
        if not file:
            return jsonify({"status": "error", "message": "No file uploaded"}), 400

        df = pd.read_excel(file, header=3)

        # Add 7 new columns into excel
        new_column_names = ['Product Level', 'Support Level', 'Squad', 'Year', 'Month', 'Aging (Days)', 'Application Category']
        for i, column_name in enumerate(new_column_names):
            df.insert(i, column_name, value=None)

        # Delete Unwanted Columns
        deleted_column_names = ['Request Number', 'Last Modified Date', 'Re-Opened Date', 'Status_Reason',
                                 'Active Incident', 'Reported Source', 'Urgency', 'Impact', 'Major Incident', 
                                 'Affected Asset', 'Model', 'Item', 'Type', 'ManufacturerName', 'SLA Due Date',
                                 'Overall Resolution Time (Second)', 'Active Resolution Time (Second)', 
                                 'Inactive Resolution Time (Second)', 'SVTTitle', 'SLACategory', 'SLMType', 
                                 'Customer Full Name', 'Customer Login ID', 'Customer Contact Number', 'Customer Corporate ID',
                                 'Customer Internet E-mail', 'Customer Company', 'Customer Cost Center', 'Customer JobTitle',
                                 'Submitter Full Name', 'Submitter Login ID', 'Internet E-mail', 'Assignee', 'Business Impact',
                                 'FCR', 'Actual Reported Date', 'Actual Resolution Date', 'Service Target Completed Date', 'Customer Department']
        df.drop(columns=deleted_column_names, errors='ignore', inplace=True)

        # Rename Columns
        rename_columns = {
            'Product Categorization Tier 2': 'Product',
            'Incident ID': 'Incident ID',
            'Service Target': 'SLA Service Target',
            'MeasurementStatus': 'SLA Measurement Status',
            'Summary': 'Incident Summary',
            'Resolution Note': 'Resolution Notes',
            'Detailed Decription': 'Detailed Description'
        }
        df.rename(columns=rename_columns, inplace=True)

        # Rearrange Columns
        rearranged_columns = [
            'Product Level', 'Support Level', 'Squad', 'Year', 'Month', 'Incident ID',
            'Product Categorization Tier 1', 'Product', 'Product Categorization Tier 3',
            'Status', 'Reported Date', 'Closed Date', 'Last Resolved Date', 'Aging (Days)',
            'Priority', 'Categorization Tier 1', 'Categorization Tier 2', 'Categorization Tier 3',
            'Incident Summary', 'SLA Service Target', 'SLA Measurement Status', 'Assigned Group',
            'Support Organization', 'Customer Location (Site)', 'Customer Location (City)',
            'Customer Location (Country)', 'Customer Organization', 'Application Category',
            'Resolution Notes', 'Cause Code', 'Detailed Description'
        ]
        df = df[rearranged_columns]

        # Input Year, Month, Support Level, Aging Day, Application Category
        df["Reported Date"] = pd.to_datetime(df['Reported Date'], errors='coerce')
        df["Last Resolved Date"] = pd.to_datetime(df['Last Resolved Date'], errors='coerce')
        df['Year'] = df['Reported Date'].dt.year
        df['Month'] = df['Reported Date'].dt.strftime('%B')
        df['Support Level'] = np.where(df['Support Organization'] == "PD UPSTREAM-DEVELOPMENT", 'Primary', 'Secondary')
        df['Aging (Days)'] = (df['Last Resolved Date'] - df['Reported Date']).dt.days
        df['Aging (Days)'] = df['Aging (Days)'].fillna("")
        df['Application Category'] = np.where(df['Product Categorization Tier 1'] == "SAP Services", 'SAP', 'Non SAP')

        # Input Product Level and Squad
        dp = pd.read_excel("Dev Tribe Incident & SR Data - Template.xlsx", sheet_name='Squad List 2', keep_default_na=False)
        df['Product'] = df['Product'].replace('NA', 'NA')
        dp['PDU Squad Scope'] = dp['PDU Squad Scope'].replace('NA', 'NA')

        for index, row in df.iterrows():
            product_category = row.get('Product', None)
            if product_category:
                corres_product_level = dp.loc[dp['CI Name'] == product_category, 'PDU Squad Scope'].values
                corres_squad = dp.loc[dp['CI Name'] == product_category, 'PDU Squad'].values
                if len(corres_product_level) > 0:
                    df.at[index, 'Product Level'] = corres_product_level[0]
                    df.at[index, 'Squad'] = corres_squad[0]

        # Create a new Excel workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write the DataFrame to the worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Define the range for the table (assuming data starts at A1)
        last_col_letter = get_column_letter(len(df.columns))
        table_range = f"A1:{last_col_letter}{len(df) + 1}"

        # Create a table
        tab = Table(displayName="Table1", ref=table_range)

        # Add a default style with striped rows
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        # Add the table to the worksheet
        ws.add_table(tab)

        # Save the workbook to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, download_name="processed_Incident.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
#endregion

#region Cleaning Work Order
@app.route('/process-excel-WorkOrder', methods=['POST'])
def process_excel_WorkOrder():

    try:
        # Get the uploaded file
        file = request.files['file']
        if not file:
            return jsonify({"status": "error", "message": "No file uploaded"}), 400

        # Import Work Order Excel file location from Download (don't forget to change \ into /)
        df = pd.read_excel(file, header=3)

        ## Methods and Operations ##

        # Add 3 new columns into excel
        new_column_names = ['Year', 'Month', 'Aging (Day)']
        for i, column_name in enumerate(new_column_names):
            df.insert(i, column_name, value=None)

        # Delete Unwanted Columns
        deleted_column_names = ['Status Reason', 'Reported Source', 'SR SLA', 'SR SLA Due Date', 
                                'Total Completion Time (SR)', 'Operational Categorization Tier 1', 
                                'Operational Categorization Tier 2', 'Operational Categorization Tier 3', 
                                'Product Categorization Tier 1', 'Product Categorization Tier 3', 'Old Mygenie Category', 
                                'Ownership', 'Customer Full Name (SR)', 'Affected Customer Full Name (WO)', 'Customer Login ID', 
                                'Customer Location (City)', 'Customer Location (Country)', 'Customer E-mail', 'Customer Company', 
                                'WO Order ID', 'WO Submit Date', 'WO Completed Date', 'WO Status', 'Detail (WO)', 'Summary(WO)', 
                                'WO Request Assignee', 'Actual Start Date', 'Customer JobTitle', 'Submitter Full Name', 'Submitter Login ID', 
                                'Total Completion Time (WO)', 'WO OLA Due Date', 'WO Measurement Status', 'WO OLA', 'WO Status Reason']
        df = df.drop(columns=deleted_column_names, errors='ignore')

        # Rename Columns
        df.rename(columns={'Product Categorization Tier 2': 'Product'}, inplace=True)
        df.rename(columns={'SR Status': 'Status'}, inplace=True)
        df.rename(columns={'SR Submit Date': 'Submit Date'}, inplace=True)
        df.rename(columns={'SR Closed Date': 'Closed Date'}, inplace=True)
        df.rename(columns={'SR Measurement Status': 'SLA Measurment'}, inplace=True)
        df.rename(columns={'Summary (SR)': 'Summary'}, inplace=True)
        df.rename(columns={'Details (SR)': 'Detail'}, inplace=True)
        df.rename(columns={'WO Assignee Group': 'Assignee Group'}, inplace=True)
        df.rename(columns={'WO Assignee Support Organization': 'Assignee Support Organization'}, inplace=True)

        # Rearrange Column
        df = df[['Year', 'Month', 'SRID', 'Product', 'Status', 'Submit Date', 'Closed Date', 
                 'Aging (Day)', 'SLA Measurment', 'Summary', 'Detail', 'Assignee Group', 
                 'Assignee Support Organization', 'Cause Code', 'Customer Department', 'Customer Cost Center',
                   'Customer Organization', 'Customer Location (Site)']]

        # Input Year, Month, Aging Day 
        df["Closed Date"] = pd.to_datetime(df['Closed Date'], errors='coerce')
        df["Submit Date"] = pd.to_datetime(df['Submit Date'], errors='coerce')
        df['Year'] = df['Submit Date'].dt.year
        df['Month'] = df['Submit Date'].dt.strftime('%B')
        df['Aging (Day)'] = (df['Closed Date'] - df['Submit Date']).dt.days
        df['Aging (Day)'] = df['Aging (Day)'].fillna("")

        # Create a new Excel workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write the DataFrame to the worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Define the range for the table (assuming data starts at A1)
        table_range = f'A1:{chr(ord("A") + len(df.columns) - 1)}{len(df) + 1}'  
        # Create a table
        tab = Table(displayName="Table1", ref=table_range)

        # Add a default style with striped rows
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        # Add the table to the worksheet
        ws.add_table(tab)

        # Save the workbook to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, download_name="processed_Incident.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
#endregion

if __name__ == '__main__':
    app.run(debug=True)
